import json
import os
import base64
import psycopg2
from psycopg2.extras import RealDictCursor

CORS_HEADERS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET, POST, DELETE, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type',
    'Access-Control-Max-Age': '86400'
}

def get_db():
    return psycopg2.connect(os.environ['DATABASE_URL'], cursor_factory=RealDictCursor)

def ok(data, status=200):
    return {'statusCode': status, 'headers': {'Content-Type': 'application/json', **CORS_HEADERS}, 'body': json.dumps(data, default=str)}

def err(msg, status=400):
    return {'statusCode': status, 'headers': {'Content-Type': 'application/json', **CORS_HEADERS}, 'body': json.dumps({'error': msg})}

def handler(event: dict, context) -> dict:
    """API для управления каталогом колесных дисков: получение, фильтрация и загрузка из Excel"""
    method = event.get('httpMethod', 'GET')

    if method == 'OPTIONS':
        return {'statusCode': 200, 'headers': CORS_HEADERS, 'body': ''}

    conn = get_db()
    cur = conn.cursor()

    try:
        if method == 'GET':
            params = event.get('queryStringParameters') or {}
            action = params.get('action', 'list')

            if action == 'filters':
                cur.execute("SELECT DISTINCT brand FROM wheel_rims WHERE brand IS NOT NULL ORDER BY brand")
                brands = [r['brand'] for r in cur.fetchall()]
                cur.execute("SELECT DISTINCT diameter FROM wheel_rims WHERE diameter IS NOT NULL ORDER BY diameter")
                diameters = [float(r['diameter']) for r in cur.fetchall()]
                cur.execute("SELECT DISTINCT pcd FROM wheel_rims WHERE pcd IS NOT NULL ORDER BY pcd")
                pcds = [r['pcd'] for r in cur.fetchall()]
                cur.execute("SELECT DISTINCT width FROM wheel_rims WHERE width IS NOT NULL ORDER BY width")
                widths = [float(r['width']) for r in cur.fetchall()]
                return ok({'brands': brands, 'diameters': diameters, 'pcds': pcds, 'widths': widths})

            conditions = []
            values = []
            if params.get('brand'):
                conditions.append("brand = %s")
                values.append(params['brand'])
            if params.get('diameter'):
                conditions.append("diameter = %s")
                values.append(float(params['diameter']))
            if params.get('pcd'):
                conditions.append("pcd = %s")
                values.append(params['pcd'])
            if params.get('width'):
                conditions.append("width = %s")
                values.append(float(params['width']))
            if params.get('in_stock') == 'true':
                conditions.append("in_stock = TRUE")

            where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''
            cur.execute(f"SELECT * FROM wheel_rims {where} ORDER BY brand, model LIMIT 200", values)
            rims = cur.fetchall()
            cur.execute(f"SELECT COUNT(*) as total FROM wheel_rims {where}", values)
            total = cur.fetchone()['total']
            return ok({'rims': rims, 'total': total})

        if method == 'POST':
            body = json.loads(event.get('body', '{}'))
            action = body.get('action', 'add')

            if action == 'upload_excel':
                file_b64 = body.get('file')
                if not file_b64:
                    return err('No file provided')

                try:
                    import openpyxl
                    from io import BytesIO
                    file_bytes = base64.b64decode(file_b64)
                    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
                    ws = wb.active

                    headers_row = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
                    col_map = {
                        'brand': next((i for i, h in enumerate(headers_row) if 'бренд' in h or 'brand' in h or 'марк' in h), None),
                        'model': next((i for i, h in enumerate(headers_row) if 'модел' in h or 'model' in h), None),
                        'diameter': next((i for i, h in enumerate(headers_row) if 'диаметр' in h or 'diameter' in h or 'r' == h), None),
                        'width': next((i for i, h in enumerate(headers_row) if 'ширин' in h or 'width' in h or 'j' == h), None),
                        'pcd': next((i for i, h in enumerate(headers_row) if 'pcd' in h or 'сверловк' in h or 'bolt' in h), None),
                        'et': next((i for i, h in enumerate(headers_row) if 'et' in h or 'вылет' in h or 'offset' in h), None),
                        'dia': next((i for i, h in enumerate(headers_row) if 'dia' in h or 'центр' in h or 'hub' in h), None),
                        'color': next((i for i, h in enumerate(headers_row) if 'цвет' in h or 'color' in h or 'colour' in h), None),
                        'material': next((i for i, h in enumerate(headers_row) if 'материал' in h or 'material' in h or 'тип' in h), None),
                        'price': next((i for i, h in enumerate(headers_row) if 'цена' in h or 'price' in h or 'стоим' in h), None),
                        'in_stock': next((i for i, h in enumerate(headers_row) if 'склад' in h or 'stock' in h or 'наличи' in h), None),
                        'image_url': next((i for i, h in enumerate(headers_row) if 'фото' in h or 'image' in h or 'url' in h or 'картинк' in h), None),
                    }

                    inserted = 0
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not any(row):
                            continue
                        def get(key):
                            idx = col_map.get(key)
                            return row[idx] if idx is not None and idx < len(row) else None

                        brand = get('brand')
                        if not brand:
                            continue

                        in_stock_val = get('in_stock')
                        if isinstance(in_stock_val, str):
                            in_stock_val = in_stock_val.lower() in ('да', 'yes', '1', 'true', '+', 'есть')
                        elif in_stock_val is None:
                            in_stock_val = True
                        else:
                            in_stock_val = bool(in_stock_val)

                        price_val = get('price')
                        try:
                            price_val = float(str(price_val).replace(' ', '').replace(',', '.')) if price_val else None
                        except:
                            price_val = None

                        cur.execute(
                            """INSERT INTO wheel_rims (brand, model, diameter, width, pcd, et, dia, color, material, price, in_stock, image_url)
                               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                            (str(brand), str(get('model') or ''), get('diameter'), get('width'),
                             str(get('pcd') or '') or None, get('et'), get('dia'),
                             str(get('color') or '') or None, str(get('material') or '') or None,
                             price_val, in_stock_val, str(get('image_url') or '') or None)
                        )
                        inserted += 1

                    conn.commit()
                    return ok({'inserted': inserted, 'message': f'Загружено {inserted} позиций'})

                except ImportError:
                    return err('openpyxl not installed', 500)

            # add single rim
            data = body
            cur.execute(
                """INSERT INTO wheel_rims (brand, model, diameter, width, pcd, et, dia, color, material, price, in_stock, image_url)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING *""",
                (data.get('brand'), data.get('model'), data.get('diameter'), data.get('width'),
                 data.get('pcd'), data.get('et'), data.get('dia'), data.get('color'),
                 data.get('material'), data.get('price'), data.get('in_stock', True), data.get('image_url'))
            )
            rim = cur.fetchone()
            conn.commit()
            return ok({'rim': rim}, 201)

        if method == 'DELETE':
            params = event.get('queryStringParameters') or {}
            rim_id = params.get('id')
            clear_all = params.get('clear_all') == 'true'

            if clear_all:
                cur.execute("DELETE FROM wheel_rims")
                conn.commit()
                return ok({'message': 'Каталог очищен'})

            if not rim_id:
                return err('Missing id')
            cur.execute("DELETE FROM wheel_rims WHERE id = %s", (rim_id,))
            conn.commit()
            return ok({'message': 'Диск удалён'})

        return err('Method not allowed', 405)

    finally:
        cur.close()
        conn.close()
